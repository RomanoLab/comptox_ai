#!/usr/bin/env python3
"""

"""

import csv
from itertools import islice
import MySQLdb
import os, sys
import pandas as pd
from pathlib import Path
from posixpath import supports_unicode_filenames
from openpyxl import load_workbook
from owlready2.class_construct import Not
from yaml import load, Loader

import owlready2
from tqdm import tqdm

import ipdb


_OWL = owlready2.get_ontology("http://www.w3.org/2002/07/owl#")


def _get_default_config_file():
    root_dir = Path(__file__).resolve().parents[2]
    if os.path.exists(os.path.join(root_dir, "CONFIG.yaml")):
        default_config_file = os.path.join(root_dir, "CONFIG.yaml")
    else:
        default_config_file = os.path.join(root_dir, "CONFIG-default.yaml")
    return default_config_file


config_file = _get_default_config_file()
with open(config_file, "r") as fp:
    CONFIG = load(fp, Loader=Loader)
DATA_DIR = CONFIG["data"]["prefix"]


def safe_add_property(entity, prop, value):
    """Add a value to a property slot on a node.

    Importantly, the method below is compatible with both functional and
    non-functional properties. If a property is functional, it either
    creates a new list or extends an existing list.

    This function cuts down on boilerplate code considerably when setting
    many property values in the ontology.
    """
    if value is None:
        return
    if _OWL.FunctionalProperty in prop.is_a:
        setattr(entity, prop._python_name, value)
    else:
        if len(getattr(entity, prop._python_name)) == 0:
            setattr(entity, prop._python_name, [value])
        else:
            if value not in getattr(entity, prop._python_name):
                getattr(entity, prop._python_name).append(value)


def get_onto_class_by_node_type(ont: owlready2.namespace.Ontology, node_label: str):
    """Get an object corresponding to an ontology class given the node label.

    `owlready2` doesn't make it easy to dynamically retrieve ontology classes.
    This uses some (relatively unsafe) string manipulation to hack together a
    solution.

    Notes
    -----
    This should be refactored if/when a better solution is available!
    """
    matches = [c for c in onto.classes() if str(c).split(".")[-1] == node_label]
    if len(matches) == 1:
        return matches[0]
    elif len(matches) == 0:
        return None
    else:
        raise ValueError(
            "Error: Something is wrong with your ontology's class hierarchy! Check for duplicate classes with '{0}' in the name".format(
                node_label
            )
        )


def safe_make_individual_name(
    indiv_name: str, indiv_class: owlready2.entity.ThingClass
):
    """Generate a (hopefully unique) descriptive name for an ontology
    individual based on the IRI column.

    Whitespaces are replaced with underscores, the class name is prepended, and
    everything is converted to lowercase.
    """
    cl = indiv_class.name.lower()
    if type(indiv_name) == str:
        nm = indiv_name.strip().replace(" ", "_").lower()
    else:
        nm = indiv_name
    return "{0}_{1}".format(cl, nm)


class DatabaseParser:
    def __init__(self, name, onto: owlready2.namespace.Ontology):
        self.name = name
        self.ont = onto

    def parse_node_type(self, node_type: str, source_node_label: str, **kwargs):
        raise NotImplementedError("Error: Base classes should override this method.")

    def _merge_node(
        self,
        fields: dict,
        node_type: str,
        source_node_label: str,
        node_properties: dict,
        merge_column: dict,
        existing_class: str = None,
    ):
        if not existing_class:
            cl = get_onto_class_by_node_type(self.ont, node_type)
        else:
            cl = get_onto_class_by_node_type(self.ont, existing_class)
        if cl == None:
            raise RuntimeError("Class label {0} not found in ComptoxAI ontology".format(node_type))

        # Find node if it exists
        merge_col = merge_column["source_column_name"]
        merge_prop_name = merge_column["data_property"].name

        # kwargs approach is necessary because keywords are only known at compile time...
        match = self.ont.search(**{merge_prop_name: fields[merge_col]})

        assert len(match) <= 1

        # Only create a new individual if we didn't find a match
        if len(match) == 0:
            individual_name = fields[source_node_label]
            new_or_matched_individual = cl(safe_make_individual_name(individual_name, cl))
        else:
            new_or_matched_individual = match[0]

            # Make sure to append the new class label if we are supposed to
            if existing_class:
                new_or_matched_individual.is_a.append(
                    get_onto_class_by_node_type(self.ont, node_type)
                )

        if node_properties:
            for field, d_prop in node_properties.items():
                # Make sure we don't overwrite the field we already matched on
                if field == merge_col:
                    continue
                safe_add_property(new_or_matched_individual, d_prop, fields[field])

    def _write_new_node(
        self,
        fields: dict,
        node_type: str,
        source_node_label: str,
        node_properties: dict = None,
    ):

        cl = get_onto_class_by_node_type(self.ont, node_type)

        individual_name = fields[source_node_label]

        # new_individual = cl(individual_name.lower())
        new_individual = cl(safe_make_individual_name(individual_name, cl))

        if node_properties:
            for field, d_prop in node_properties.items():
                try:
                    safe_add_property(new_individual, d_prop, fields[field])
                except KeyError:
                    # TODO: Log missing properties!
                    pass


class FlatFileDatabaseParser(DatabaseParser):
    def __init__(self, name: str, destination: owlready2.namespace.Ontology):
        super().__init__(name, destination)

    def get_file_pointer_by_flatfile_name(self, filename):
        file_pointer = open(os.path.join(DATA_DIR, self.name, filename), "r", encoding="utf-8", errors="ignore")
        return file_pointer

    def get_file_pointer_by_node_label(self, node_label):
        raise NotImplementedError()

    def _store_table_headers(self, header_data, delimiter=","):
        headers = header_data.split(delimiter)

        print("FILE_HEADERS:")
        for h in headers:
            print(h)

        self.headers = headers

    def parse_relationship_type(self, relationship_type, source_filename: str, fmt: str, parse_config: dict, inverse_relationship_type = None, merge: bool = False, skip: bool = False):
        if skip:
            return

        print("PARSING RELATIONSHIP TYPE: {0}".format(relationship_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        s_node_type = parse_config["subject_node_type"]
        o_node_type = parse_config["object_node_type"]

        sub_match_prop_name = parse_config["subject_match_property"].name
        obj_match_prop_name = parse_config["object_match_property"].name

        if not fmt == "xlsx":
            fp = self.get_file_pointer_by_flatfile_name(source_filename)

        if fmt == "csv":
            reader = csv.reader(fp)

            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            node_iter = reader

        elif fmt == "tsv":
            reader = csv.reader(fp, delimiter="\t")

            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            node_iter = reader

        else:
            raise RuntimeError("Unsupported flatfile format: {0}".format(format))

        compound_fields = None
        if "compound_fields" in parse_config:
            compound_fields = list(parse_config["compound_fields"].keys())

        data_transforms = None
        if "data_transforms" in parse_config:
            data_transforms = list(parse_config["data_transforms"].keys())

        for rel in tqdm(node_iter):
            fields = dict(zip(headers, rel))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if compound_fields:
                for cf in compound_fields:
                    cf_config = parse_config["compound_fields"][cf]
                    cf_data = fields[cf].split(cf_config["delimiter"])
                    for subfield in cf_data:
                        s = subfield.split(cf_config["field_split_prefix"])
                        fields[s[0]] = s[-1]

            if data_transforms:
                for dt in data_transforms:
                    fields[dt] = parse_config["data_transforms"][dt](fields[dt])

            subj_ids = fields[parse_config["subject_column_name"]]
            obj_ids = fields[parse_config["object_column_name"]]

            if type(subj_ids) is not list:
                subj_ids = [subj_ids]
            if type(obj_ids) is not list:
                obj_ids = [obj_ids]

            for sid in subj_ids:
                for oid in obj_ids:
                    subject_match = self.ont.search(**{sub_match_prop_name: sid})
                    if len(subject_match) == 0:
                        continue
                    try:
                        assert len(subject_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()
                    
                    object_match = self.ont.search(**{obj_match_prop_name: oid})
                    if len(object_match) == 0:
                        continue
                    try:
                        assert len(object_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()

                    safe_add_property(subject_match[0], relationship_type, object_match[0])
                    if inverse_relationship_type:
                        safe_add_property(object_match[0], inverse_relationship_type, subject_match[0])

    def parse_node_type(
        self,
        node_type: str,
        source_filename: str,
        fmt: str,
        parse_config: dict,
        merge: bool = True,
        append_class: bool = False,
        existing_class: str = None,
        skip: bool = False,
    ):
        """
    Parameters
    ----------
    node_type : str
      String name corresponding to the node type in the destination database.
    source_node_label : str
      String used to identify the node type in the source database.
    merge : bool
      When `True`, entities that already exist in the database will be *merged*
      with the new data. When `False`, a new entity will be created in the
      graph database regardless of whether a matching entity already exists in
      the database.
    """
        if skip:
            return

        print("PARSING NODE TYPE: {0}".format(node_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        ont_class = get_onto_class_by_node_type(self.ont, node_type)

        if not fmt == "xlsx":
            fp = self.get_file_pointer_by_flatfile_name(source_filename)

        # Use the specified file format to create the following two items:
        # - `headers` - a list of column headers
        # - `node_iter` - an iterable that returns individual node records from the file
        #
        # Note: Each item returned by `node_iter` must be the same length as `headers`

        if fmt == "csv":
            reader = csv.reader(fp)

            if not parse_config["headers"]:
                raise NotImplementedError(
                    "Parsing of files without column headers is not currently supported"
                )
            if type(parse_config["headers"]) == list:
                headers = parse_config["headers"]
            elif parse_config["headers"] == True:
                headers = next(reader)

            if "skip_n_lines" in parse_config:
                for _ in range(parse_config["skip_n_lines"]):
                    next(reader)

            # for line in tqdm(reader):
            #     data = dict(zip(headers, line))
            node_iter = reader

        elif fmt == "tsv":
            reader = csv.reader(fp, delimiter="\t")

            if not parse_config["headers"]:
                raise NotImplementedError(
                    "Parsing of files without column headers is not currently supported"
                )
            headers = next(reader)

            # for line in tqdm(reader):
            #     data = dict(zip(headers, line))
            node_iter = reader

        elif fmt == "xlsx":
            sfn = os.path.join(DATA_DIR, self.name, source_filename)
            print("LOADING .xlsx FILE - PLEASE BE PATIENT...")
            wb = load_workbook(filename=sfn, read_only=True, data_only=True)
            print("...done.")

            # Hopefully we only have one worksheet!
            ws = wb[wb.sheetnames[0]]

            header_cells = ws[1]
            headers = [hc.value for hc in header_cells]

            # for line in tqdm(ws.iter_rows(min_row=2)):
            #     vals = [l.value for l in line]
            #     data = dict(zip(headers, vals))
            node_iter = ws.iter_rows(min_row=2)

        elif fmt == "tsv-pandas":
            df = pd.read_csv(
                os.path.join(DATA_DIR, self.name, source_filename), sep="\t"
            )

            headers = list(df.columns)

            # Use a generator expression to coerce the Pandas iterator into our desired format
            node_iter = (x for _, x in df.iterrows())

        else:
            raise RuntimeError("Unsupported flatfile format: {0}".format(format))

        compound_fields = None
        if "compound_fields" in parse_config:
            compound_fields = list(parse_config["compound_fields"].keys())

        data_transforms = None
        if "data_transforms" in parse_config:
            data_transforms = list(parse_config["data_transforms"].keys())

        if merge:
            merge_column = parse_config["merge_column"]
            if append_class:
                existing_class_label = existing_class
            else:
                existing_class_label = None

        for node in tqdm(node_iter):
            fields = dict(zip(headers, node))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if compound_fields:
                for cf in compound_fields:
                    cf_config = parse_config["compound_fields"][cf]
                    cf_data = fields[cf].split(cf_config["delimiter"])
                    for subfield in cf_data:
                        s = subfield.split(cf_config["field_split_prefix"])
                        fields[s[0]] = s[-1]

            if data_transforms:
                for dt in data_transforms:
                    fields[dt] = parse_config["data_transforms"][dt](fields[dt])

            if merge:
                self._merge_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                    merge_column,
                    existing_class_label,
                )
            else:
                self._write_new_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                )

        fp.close()


class MySQLDatabaseParser(DatabaseParser):
    def __init__(
        self, name: str, destination: owlready2.namespace.Ontology, config_dict: dict
    ):
        super().__init__(name, destination)

        self.conn = MySQLdb.Connection(
            host=config_dict["host"],
            user=config_dict["user"],
            passwd=config_dict["passwd"],
            db=name,
        )

    def parse_relationship_type(self, relationship_type, parse_config: dict, inverse_relationship_type = None, merge: bool = False, skip: bool = False):
        print("PARSING RELATIONSHIP TYPE: {0}".format(relationship_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))
        
        s_node_type = parse_config["subject_node_type"]
        o_node_type = parse_config["object_node_type"]

        c = self.conn.cursor()

        if "custom_sql_query" in parse_config:
            fetch_lines_query = parse_config["custom_sql_query"]
        else:
            fetch_lines_query = "SELECT * FROM {0};".format(parse_config["source_table"])
        c.execute(fetch_lines_query)

        headers = [cc[0] for cc in c.description]

        sub_match_prop_name = parse_config["subject_match_property"].name
        obj_match_prop_name = parse_config["object_match_property"].name

        for _ in tqdm(range(c.rowcount)):
            row = c.fetchone()

            fields = dict(zip(headers, row))

            if "compound_fields" in parse_config:
                for c_f_name, c_f_config in parse_config["compound_fields"].items():
                    fields[c_f_name] = fields[c_f_name].split(c_f_config["delimiter"])

            if "data_transforms" in parse_config:
                for dt_f_name, dt_f_transform in parse_config["data_transforms"].items():
                    if type(fields[dt_f_name]) == list:
                        fields[dt_f_name] = [dt_f_transform(x) for x in fields[dt_f_name]]
                    else:
                        fields[dt_f_name] = dt_f_transform(fields[dt_f_name])

            # Get subject and object nodes
            if parse_config["source_table_type"] == "foreignKey":
                subj_ids = fields[parse_config["subject_column_name"]]
                obj_ids = fields[parse_config["object_column_name"]]

                if type(subj_ids) is not list:
                    subj_ids = [subj_ids]
                if type(obj_ids) is not list:
                    obj_ids = [obj_ids]
            else:
                raise NotImplementedError("`source_table_type` not recognized")

            for sid in subj_ids:
                for oid in obj_ids:
                    subject_match = self.ont.search(**{sub_match_prop_name: sid})
                    if len(subject_match) == 0:
                        continue
                    try:
                        assert len(subject_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()
                    object_match = self.ont.search(**{obj_match_prop_name: oid})
                    if len(object_match) == 0:
                        continue
                    try:
                        assert len(object_match) == 1
                    except AssertionError:
                        ipdb.set_trace()
                        print()

                    safe_add_property(subject_match[0], relationship_type, object_match[0])
                    if inverse_relationship_type:
                        safe_add_property(object_match[0], inverse_relationship_type, subject_match[0])

        c.close()

    def parse_node_type(
        self,
        node_type: str,
        source_table: str,
        parse_config: dict,
        merge: bool = True,
        append_class: bool = False,
        existing_class: str = None,
        skip: bool = False,
    ):
        if skip:
            return

        print("PARSING NODE TYPE: {0}".format(node_type))
        print("FROM SOURCE DATABASE: {0}".format(self.name))

        c = self.conn.cursor()

        if "custom_sql_query" in parse_config:
            fetch_lines_query = parse_config["custom_sql_query"]
        else:
            fetch_lines_query = "SELECT * FROM {0};".format(source_table)
        c.execute(fetch_lines_query)

        # Get column names
        headers = [cc[0] for cc in c.description]

        if merge:
            merge_column = parse_config["merge_column"]
            if append_class:
                existing_class_label = existing_class
            else:
                existing_class_label = None

        for _ in tqdm(range(c.rowcount)):  # would `while` be safer?
            row = c.fetchone()

            fields = dict(zip(headers, row))

            if "filter_column" in parse_config:
                # Do a simple string matching. This isn't 'safe' and may lead
                # to issues down the line, but for now we are using it cautiously.
                # TODO: Refactor per above note
                if (
                    parse_config["filter_value"]
                    not in fields[parse_config["filter_column"]]
                ):
                    continue

            if merge:
                self._merge_node(
                    fields=fields,
                    node_type=node_type,
                    source_node_label=parse_config["iri_column_name"],
                    node_properties=parse_config["data_property_map"],
                    merge_column=merge_column,
                    existing_class=existing_class_label,
                )
            else:
                self._write_new_node(
                    fields,
                    node_type,
                    parse_config["iri_column_name"],
                    parse_config["data_property_map"],
                )

        c.close()


def print_ontology_stats(ont: owlready2.namespace.Ontology):

    num_classes = sum(1 for _ in ont.classes())
    num_individuals = sum(1 for _ in ont.individuals())
    num_dps = sum(1 for _ in ont.data_properties())
    num_ops = sum(1 for _ in ont.object_properties())

    print("  ===ONTOLOGY STATISTICS===")
    print()

    print("  Number of classes:           {0}".format(num_classes))
    print("  Number of data properties:   {0}".format(num_dps))
    print("  Number of object properties: {0}".format(num_ops))
    print()
    print("  Number of individuals:       {0}".format(num_individuals))
    print("  Number of relations:         {0}".format("[UNIMPLEMENTED]"))
    print()


if __name__ == "__main__":
    #onto = owlready2.get_ontology("file://D:\\projects\\comptox_ai\\comptox.rdf").load()
    onto = owlready2.get_ontology("file://D:\\projects\\comptox_ai\\comptox_mid.rdf").load()

    # open config file:
    with open("../../CONFIG.yaml", 'r') as fp:
        cnf = yaml.load(fp)
    mysql_config = dict()
    mysql_config["host"] = cnf["mysql"]["host"]
    mysql_config["user"] = cnf["mysql"]["user"]
    mysql_config["passwd"] = cnf["mysql"]["passwd"]

    #print("INITIAL ONTOLOGY STATISTICS:")
    #print_ontology_stats(onto)

    epa = FlatFileDatabaseParser("epa", onto)
    ncbigene = FlatFileDatabaseParser("ncbigene", onto)
    #ctd = FlatFileDatabaseParser("ctd", onto)
    drugbank = FlatFileDatabaseParser("drugbank", onto)
    hetionet = FlatFileDatabaseParser("hetionet", onto)
    aopdb = MySQLDatabaseParser("aopdb", onto, mysql_config)

    # Add nodes and node properties in a carefully specified order

    #####################
    # EPA COMPTOX NODES #
    #####################
    epa.parse_node_type(
        node_type="Chemical",
        source_filename="PubChem_DTXSID_mapping_file.txt",
        fmt="tsv",
        parse_config={
            "iri_column_name": "DTXSID",
            "headers": True,
            "data_property_map": {
                "CID": onto.xrefPubchemCID,
                "SID": onto.xrefPubchemSID,
                "DTXSID": onto.xrefDTXSID,
            },
        },
        merge=False,
        skip=True,
    )
    epa.parse_node_type(
        node_type="Chemical",
        source_filename="Dsstox_CAS_number_name.csv",
        fmt="csv",
        parse_config={
            "iri_column_name": "dsstox_substance_id",
            "headers": True,
            "data_property_map": {
                "casrn": onto.xrefCasRN,
                "preferred_name": onto.commonName,
                "dsstox_substance_id": onto.xrefDTXSID,
            },
            "merge_column": {
                "source_column_name": "dsstox_substance_id",
                "data_property": onto.xrefDTXSID,
            },
        },
        merge=True,
        skip=True,
    )

    # ipdb.set_trace()
    # #############
    # # CTD NODES #
    # #############
    # ctd.parse_node_type(
    #     node_type="Chemical",
    #     source_filename="CTD_chemicals.csv",
    #     fmt="csv",
    #     parse_config={
    #         "iri_column_name": "CasRN",
    #         "headers": ["ChemicalName", "ChemicalID", "CasRN", "Definition", "ParentIDs", "TreeNumbers", "ParentTreeNumbers", "Synonyms"],
    #         "skip_n_lines": 29,
    #         "data_property_map": {
    #             "CasRN": onto.xrefCasRN,
    #             "ChemicalID": onto.xrefMeSHUI,
    #         },
    #         "merge_column": {
    #             "source_column_name": "CasRN",
    #             "data_property": onto.xrefCasRN
    #         },
    #         "data_transforms": {
    #             "ChemicalID": lambda x: x.split(":")[-1]
    #         }
    #     },
    #     merge=True,
    #     skip=False
    # )

    ##################
    # DRUGBANK NODES #
    ##################
    drugbank.parse_node_type(
        node_type="Chemical",
        source_filename="drug_links.csv",
        fmt="csv",
        parse_config={
            "iri_column_name": "DrugBank ID",
            "headers": True,
            "data_property_map": {
                "DrugBank ID": onto.xrefDrugbank,
                "CAS Number": onto.xrefCasRN
            },
            "merge_column": {
                "source_column_name": "CAS Number",
                "data_property": onto.xrefCasRN,
            },
        },
        merge=True,
        skip=True
    )

    ###################
    # NCBI GENE NODES #
    ###################
    ncbigene.parse_node_type(
        node_type="Gene",
        source_filename="Homo_sapiens.gene_info",
        fmt="tsv-pandas",
        parse_config={
            "compound_fields": {
                "dbXrefs": {"delimiter": "|", "field_split_prefix": ":"}
            },
            "iri_column_name": "Symbol",
            "headers": True,
            "data_property_map": {
                "GeneID": onto.xrefNcbiGene,
                "Symbol": onto.geneSymbol,
                "type_of_gene": onto.typeOfGene,
                "MIM": onto.xrefOMIM,
                "HGNC": onto.xrefHGNC,
                "Ensembl": onto.xrefEnsembl,
                # TODO: Parse Feature_type and other columns
            },
        },
        merge=False,
        skip=True,
    )

    ################
    # AOP-DB NODES #
    ################
    aopdb.parse_node_type(
        node_type="AOP",
        source_table="aop_info",
        parse_config={
            "iri_column_name": "AOP_id",
            "data_property_map": {
                "AOP_name": onto.aopName,
                "AOP_id": onto.xrefAOPWikiAOPID,
            },
        },
        merge=False,
        skip=True,
    )
    aopdb.parse_node_type(
        node_type="KeyEvent",
        source_table="event_info",
        parse_config={
            "iri_column_name": "event_id",
            "data_property_map": {
                "event_name": onto.keyEventName,
                "event_id": onto.xrefAOPWikiKEID,
            },
            "data_transforms": {
                "name": lambda x: x.split("; ")[0].strip()  # For some reason, a lot of the event names are duplicated
            },
        },
        merge=False,
        skip=True,
    )
    aopdb.parse_node_type(
        node_type="MolecularInitiatingEvent",
        source_table="event_info",
        parse_config={
            "iri_column_name": "event_id",
            "data_property_map": {"event_id": onto.xrefAOPWikiKEID},
            "filter_column": "event_type",
            "filter_value": "molecular-initiating-event",
            "merge_column": {
                "source_column_name": "event_id",
                "data_property": onto.xrefAOPWikiKEID,
            },
        },
        merge=True,
        append_class=True,
        existing_class="KeyEvent",
        skip=True,
    )
    aopdb.parse_node_type(
        node_type="AdverseOutcome",
        source_table="event_info",
        parse_config={
            "iri_column_name": "event_id",
            "data_property_map": {"event_id": onto.xrefAOPWikiKEID},
            "filter_column": "event_type",
            "filter_value": "adverse-outcome",
            "merge_column": {
                "source_column_name": "event_id",
                "data_property": onto.xrefAOPWikiKEID,
            },
        },
        merge=True,
        append_class=True,
        existing_class="KeyEvent",
        skip=True,
    )
    # To denote something as a stressor we just add the stressor ID to a
    # Chemical node. 
    # TODO: Make sure to link KeyEvent nodes to Chemicals
    aopdb.parse_node_type(
        node_type="Chemical",
        source_table="stressor_info",
        parse_config={
            "iri_column_name": "DTX_id",
            "data_property_map": {"stressor_id": onto.xrefAOPWikiStressorID},
            "merge_column": {
                "source_column_name": "DTX_id",
                "data_property": onto.xrefDTXSID,
            },
        },
        merge=True,
        skip=True
    )
    aopdb.parse_node_type(
        node_type="Pathway",
        source_table="stressor_info",
        parse_config={
            "iri_column_name": "path_id",
            "data_property_map": {
                "path_id": onto.pathwayId,
                "path_name": onto.commonName,
                "ext_source": onto.sourceDatabase,
            },
            "custom_sql_query": "SELECT DISTINCT path_id, path_name, ext_source FROM aopdb.pathway_gene WHERE tax_id = 9606;"
        },
        merge=False,
        skip=True
    )

    # with open("D:\\projects\\comptox_ai\\comptox_mid.rdf", "wb") as fp:
    #     onto.save(file=fp, format="rdfxml")
    #ipdb.set_trace()

    # Add relationships and relationship properties
    # Options for "source_type":
    #  - MySQL source:
    #      - `join_table`: Each row in a table represents a relationship
    #        between two entities each defined in their own tables
    #      - `foreign_key`: Starting from the subject node, a relationship
    #        is inferred by looking up the object node based on a foreign
    #        key value in the subject's source table

    ########################
    # AOP-DB RELATIONSHIPS #
    ########################
    aopdb.parse_relationship_type(
        relationship_type=onto.keIncludedInAOP,
        inverse_relationship_type=onto.aopIncludesKE,
        parse_config = {
            "subject_node_type": onto.KeyEvent,
            "subject_column_name": "event_id",
            "subject_match_property": onto.xrefAOPWikiKEID,
            "object_node_type": onto.AOP,
            "object_column_name": "AOP_ids",
            "object_match_property": onto.xrefAOPWikiAOPID,
            "source_table_type": "foreignKey",
            "source_table": "event_info",
            "compound_fields": {
                "AOP_ids": {"delimiter": "; "}  # Notice the space after the ";"
            },
            "data_transforms": {
                "AOP_ids": lambda x: int(x)
            },
        },
        merge=False,
        skip=False
    )
    aopdb.parse_relationship_type(
        relationship_type=onto.keyEventTriggeredBy,
        parse_config = {
            "subject_node_type": onto.KeyEvent,
            "subject_column_name": "event_id",
            "subject_match_property": onto.xrefAOPWikiKEID,
            "object_node_type": onto.Chemical,
            "object_column_name": "DTX_id",
            "object_match_property": onto.xrefDTXSID,
            "source_table_type": "foreignKey",
            "source_table": "aop_stressor",
        },
        merge=False,
        skip=False
    )
    aopdb.parse_relationship_type(
        relationship_type=onto.geneInPathway,
        inverse_relationship_type=onto.PathwayContainsGene,
        parse_config = {
            "subject_node_type": onto.Gene,
            "subject_column_name": "entrez",
            "subject_match_property": onto.xrefNcbiGene,
            "object_node_type": onto.Pathway,
            "object_column_name": "path_id",
            "object_match_property": onto.pathwayId,
            "custom_sql_query": "SELECT * FROM aopdb.pathway_gene WHERE tax_id = 9606;",
            "source_table_type": "foreignKey",
            "source_table": "pathway_gene",
        },
        merge=False,
        skip=False
    )

    #####################
    # CTD RELATIONSHIPS #
    #####################
    # ctd.parse_relationship_type(
    #     relationship_type=onto.chemicalIncreasesExpression,
    #     source_filename="CTD_chem_gene_ixns.csv",
    #     fmt="csv",
    #     parse_config = {
    #         "subject_node_type": onto.Chemical,
    #         "subject_column_name": "ChemicalID",
    #         "subject_match_property": onto.xrefMeSHUI,
    #         "object_node_type": onto.Gene,
    #         "object_column_name": "GeneID",
    #         "object_match_property": onto.xrefNcbiGene,
    #         "source_filename": "CTD_chem_gene_ixns.csv",
    #         "headers": ["ChemicalName","ChemicalID","CasRN","GeneSymbol","GeneID","GeneForms","Organism","OrganismID","Interaction","InteractionActions","PubMedIDs"],
    #         "skip_n_lines": 29,
    #         "filter_column": "InteractionActions",
    #         "filter_value": "increases^expression",
    #     },
    #     merge=False,
    #     skip=False
    # )
    # ctd.parse_relationship_type(
    #     relationship_type=onto.chemicalDecreasesExpression,
    #     source_filename="CTD_chem_gene_ixns.csv",
    #     fmt="csv",
    #     parse_config = {
    #         "subject_node_type": onto.Chemical,
    #         "subject_column_name": "ChemicalID",
    #         "subject_match_property": onto.xrefMeSHUI,
    #         "object_node_type": onto.Gene,
    #         "object_column_name": "GeneID",
    #         "object_match_property": onto.xrefNcbiGene,
    #         "source_filename": "CTD_chem_gene_ixns.csv",
    #         "headers": ["ChemicalName","ChemicalID","CasRN","GeneSymbol","GeneID","GeneForms","Organism","OrganismID","Interaction","InteractionActions","PubMedIDs"],
    #         "skip_n_lines": 29,
    #         "filter_column": "InteractionActions",
    #         "filter_value": "decreases^expression",
    #     },
    #     merge=False,
    #     skip=False
    # )

    ##########################
    # HETIONET RELATIONSHIPS #
    ##########################
    hetionet.parse_relationship_type(
        relationship_type=onto.chemicalIncreasesExpression,
        source_filename="hetionet-v1.0-edges.sif",
        fmt="tsv",
        parse_config={
            "subject_node_type": onto.Chemical,
            "subject_column_name": "source",
            "subject_match_property": onto.xrefDrugbank,
            "object_node_type": onto.Gene,
            "object_column_name": "target",
            "object_match_property": onto.xrefNcbiGene,
            "filter_column": "metaedge",
            "filter_value": "CuG",
            "headers": True,
            "data_transforms": {
                "source": lambda x: x.split("::")[-1],
                "target": lambda x: int(x.split("::")[-1])  # I foresee this causing problems in the future - should all IDs be cast to str?
            },
        },
        merge=False,
        skip=False
    )
    hetionet.parse_relationship_type(
        relationship_type=onto.chemicalDecreasesExpression,
        source_filename="hetionet-v1.0-edges.sif",
        fmt="tsv",
        parse_config={
            "subject_node_type": onto.Chemical,
            "subject_column_name": "source",
            "subject_match_property": onto.xrefDrugbank,
            "object_node_type": onto.Gene,
            "object_column_name": "target",
            "object_match_property": onto.xrefNcbiGene,
            "filter_column": "metaedge",
            "filter_value": "CdG",
            "headers": True,
            "data_transforms": {
                "source": lambda x: x.split("::")[-1],
                "target": lambda x: int(x.split("::")[-1])  # I foresee this causing problems in the future - should all IDs be cast to str?
            },
        },
        merge=False,
        skip=False
    ),
    hetionet.parse_relationship_type(
        relationship_type=onto.chemicalBindsGene,
        source_filename="hetionet-v1.0-edges.sif",
        fmt="tsv",
        parse_config={
            "subject_node_type": onto.Chemical,
            "subject_column_name": "source",
            "subject_match_property": onto.xrefDrugbank,
            "object_node_type": onto.Gene,
            "object_column_name": "target",
            "object_match_property": onto.xrefNcbiGene,
            "filter_column": "metaedge",
            "filter_value": "CbG",
            "headers": True,
            "data_transforms": {
                "source": lambda x: x.split("::")[-1],
                "target": lambda x: int(x.split("::")[-1])  # I foresee this causing problems in the future - should all IDs be cast to str?
            },
        },
        merge=False,
        skip=False
    )
    hetionet.parse_relationship_type(
        relationship_type=onto.geneInteractsWithGene,
        source_filename="hetionet-v1.0-edges.sif",
        fmt="tsv",
        parse_config={
            "subject_node_type": onto.Gene,
            "subject_column_name": "source",
            "subject_match_property": onto.xrefNcbiGene,
            "object_node_type": onto.Gene,
            "object_column_name": "target",
            "object_match_property": onto.xrefNcbiGene,
            "filter_column": "metaedge",
            "filter_value": "GiG",
            "headers": True,
            "data_transforms": {
                "source": lambda x: x.split("::")[-1],
                "target": lambda x: int(x.split("::")[-1])  # I foresee this causing problems in the future - should all IDs be cast to str?
            },
        },
        merge=False,
        skip=False
    )

    ipdb.set_trace()

    # Save the ontology to a new file
    with open("D:\\projects\\comptox_ai\\comptox_populated.rdf", "wb") as fp:
        onto.save(file=fp, format="rdfxml")
